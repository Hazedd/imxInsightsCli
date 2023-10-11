import re
from pathlib import Path
from typing import Optional, List

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


def replace_counter_in_path(path):
    pattern = r"\.\d+\."
    return re.sub(pattern, ".", path)


def get_excel_last_column(sheet: Worksheet):
    for max_row, row in enumerate(sheet, 1):
        return row[-1]


def excel_remove_autofilter(sheet: Worksheet):
    if sheet.auto_filter:
        sheet.auto_filter.ref = None


def excel_set_autofilter(sheet: Worksheet, row: int, column_letter_start: str, column_letter_end: str):
    sheet.auto_filter.ref = f"{column_letter_start}{row}:{column_letter_end}{row}"


class ExcelScope:
    def __init__(self, scope_excel_file_path: Path):
        self.df_scope: Optional[pd.DataFrame] = None
        self.scope_version: Optional[str] = None
        self._scope_excel_file_path = scope_excel_file_path
        self.set_scope_data(scope_excel_file_path)

        self.scope_types = [item for item in self.df_scope.columns.tolist() if item not in ["path", "target"]]
        self.scope_dict = self.df_scope.to_dict("index")

    def __repr__(self):
        return f"{self._scope_excel_file_path}, scopes: {', '.join([item for item in self.scope_types])}"

    def set_scope_data(self, scope_excel_file_path: Path):
        df_info_ = pd.read_excel(scope_excel_file_path, sheet_name="info", index_col=0)
        self.scope_version = scope_version = df_info_.loc["uitwisselscope_version:"][0]

        df_scope_ = pd.read_excel(scope_excel_file_path, sheet_name="as-dataset", index_col=0)
        df_scope_["index"] = df_scope_["path"] + "/" + df_scope_["target"].replace({".gml:": "."}, regex=True)
        df_scope_.set_index(["index"], inplace=True)
        self.df_scope = df_scope_

    @staticmethod
    def set_scope_info(scope_excel_file_path: Path):
        pass


    def filter_keys(self, key_suffix: str):
        return [
            value for key, value in self.scope_dict.items() if
            key.endswith(key_suffix) and
            key[-len(key_suffix)-1:-len(key_suffix)] in [".", ""]
        ]

    def add_scope(self, diff_excel_file_path: Path, out_excel_file_path: Path, scope_types: Optional[List[str]] = None):
        fill_scoped = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

        def build_header_get_header_row():
            for _ in scope_types:
                sheet.insert_rows(1)
            for idx_, item in enumerate(scope_types):
                sheet.cell(row=idx_ + 1, column=1, value=item)
            return len(scope_types) + 1

        if scope_types is None:
            scope_types = [item for item in self.scope_types]
        else:
            scope_types = [item for item in self.scope_types if item in scope_types]

        workbook = openpyxl.load_workbook(diff_excel_file_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            if sheet['A1'].value == "index" and sheet['B1'].value == "puic" and sheet['C1'].value == "path":
                path = sheet['C2'].value
                excel_remove_autofilter(sheet)
                last_column_in_header = get_excel_last_column(sheet)
                header_row = build_header_get_header_row()
                excel_set_autofilter(sheet, header_row, "A", last_column_in_header.column_letter)

                for cell in sheet[header_row]:
                    key_suffix = replace_counter_in_path(f"{path}/{cell.value}")
                    filtered = self.filter_keys(key_suffix)
                    assert len(filtered) < 2, "should return 1 or less"

                    if len(filtered) != 0:
                        for idx, scope in enumerate(scope_types):
                            if scope in filtered[0].keys():
                                sheet.cell(row=idx + 1, column=cell.column, value=True if filtered[0][scope] == 1 else False)
                                cell.fill = fill_scoped

        workbook.save(out_excel_file_path)


if __name__ == "__main__":
    pass
    # excel_scope = ExcelScope(Path("otl-overview-Situation-5_0_0-v3.xlsx"))
    # print(excel_scope)
    # excel_scope.add_scope(Path("diff.xlsx"), Path("your_updated_workbook.xlsx"), ["Is van toepassing", "Invullen door ingenieursbureau"])

